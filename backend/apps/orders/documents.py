from __future__ import annotations

import logging
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import TYPE_CHECKING, Any

from django.core.exceptions import ObjectDoesNotExist
from django.core.files.base import ContentFile
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from apps.orders.models import Order, OrderDocument

logger = logging.getLogger(__name__)

TYPE_INVOICE = 'invoice'
TYPE_TTN = 'ttn'
TYPE_CMR = 'cmr'
TYPE_ACT = 'act'

DOC_PREFIX = {
    TYPE_INVOICE: 'SF',
    TYPE_TTN: 'TTN',
    TYPE_CMR: 'CMR',
    TYPE_ACT: 'AKT',
}

DOC_TITLES = {
    TYPE_INVOICE: 'Hisob-faktura',
    TYPE_TTN: 'Tovar-transport yuk xati (TTN)',
    TYPE_CMR: 'CMR — xalqaro yuk xati',
    TYPE_ACT: 'Bajarilgan ishlar dalolatnomasi',
}


def _dec(value) -> Decimal:
    try:
        return Decimal(str(value or 0))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal('0')


def _money(value) -> str:
    amount = _dec(value)
    formatted = f'{amount:,.2f}'.replace(',', ' ')
    return formatted


def _dt(value) -> str:
    if not value:
        return ''
    if timezone.is_aware(value):
        value = timezone.localtime(value)
    return value.strftime('%d.%m.%Y %H:%M')


def _date(value) -> str:
    if not value:
        return ''
    if timezone.is_aware(value):
        value = timezone.localtime(value)
    return value.strftime('%d.%m.%Y')


def _person(user) -> dict[str, str]:
    if not user:
        return {'name': '', 'phone': '', 'inn': ''}
    name = f'{user.first_name or ""} {user.last_name or ""}'.strip() or (user.phone or '')
    return {
        'name': name,
        'phone': user.phone or '',
        'inn': getattr(user, 'company_inn', None) or '',
    }


def _company_party(user, *, company=None, fallback_name='') -> dict[str, str]:
    person = _person(user)
    company_name = ''
    address = phone = director = bank_name = bank_account = mfo = oked = ''
    inn = person['inn']
    if company is None and inn:
        from apps.users.models import Company
        company = Company.objects.filter(inn=inn).first()
    if company:
        company_name = company.name or ''
        address = getattr(company, 'address', '') or ''
        phone = getattr(company, 'phone', '') or person['phone']
        director = getattr(company, 'director_name', '') or ''
        bank_name = getattr(company, 'bank_name', '') or ''
        bank_account = getattr(company, 'bank_account', '') or ''
        mfo = getattr(company, 'mfo', '') or ''
        oked = getattr(company, 'oked', '') or ''
        inn = company.inn or inn
    return {
        'name': company_name or fallback_name or person['name'],
        'inn': inn,
        'address': address,
        'phone': phone or person['phone'],
        'director': director,
        'bank_name': bank_name,
        'bank_account': bank_account,
        'mfo': mfo,
        'oked': oked,
    }


def _city_name(city) -> str:
    if not city:
        return ''
    return city.name_uz or city.name_ru or city.name_en or ''


def document_number(doc_type: str, order_id: int) -> str:
    prefix = DOC_PREFIX.get(doc_type, 'DOC')
    return f'{prefix}-{int(order_id):06d}'


def build_document_snapshot(order: Order) -> dict[str, Any]:
    from apps.payments.escrow import commission_percent
    from apps.orders.amount_words import amount_in_words_uz

    ad = order.advertisement
    client = _person(order.client)
    driver = _person(order.driver)
    shipper = _company_party(order.client)
    carrier = _company_party(
        order.driver,
        fallback_name=driver['name'],
    )
    dep_city = getattr(ad, 'departure_city', None)
    dest_city = getattr(ad, 'destination_city', None)
    dep_country = getattr(getattr(dep_city, 'country', None), 'name_uz', '') or getattr(getattr(dep_city, 'country', None), 'code', '') or ''
    dest_country = getattr(getattr(dest_city, 'country', None), 'name_uz', '') or getattr(getattr(dest_city, 'country', None), 'code', '') or ''
    pickup_full = ', '.join(part for part in (_city_name(dep_city), getattr(ad, 'departure_address', '') or '') if part)
    delivery_full = ', '.join(part for part in (_city_name(dest_city), getattr(ad, 'destination_address', '') or '') if part)
    route = ' — '.join(part for part in (_city_name(dep_city), _city_name(dest_city)) if part)
    receiver_name = getattr(ad, 'receiver_name', '') or ''
    consignee = {
        'name': receiver_name or shipper['name'],
        'inn': shipper['inn'] if not receiver_name else '',
        'address': delivery_full,
        'phone': getattr(ad, 'receiver_phone', '') or shipper['phone'],
        'director': '',
        'bank_name': '',
        'bank_account': '',
        'mfo': '',
        'oked': '',
    }
    vehicle = (
        order.driver.vehicles.order_by('-is_verified', '-id').first()
        if getattr(order, 'driver_id', None)
        else None
    )
    stops = []
    for stop in order.route_stops.all().order_by('sequence'):
        stops.append({
            'sequence': stop.sequence,
            'stop_type': stop.stop_type,
            'label': stop.label or '',
            'address': stop.address or '',
            'status': stop.status,
            'completed_at': _dt(stop.completed_at),
        })
    pod = None
    try:
        pod = order.proof_of_delivery
    except ObjectDoesNotExist:
        pod = None
    amount = _dec(order.total_amount)
    commission = Decimal('0')
    driver_net = amount
    escrow_status = ''
    try:
        escrow = order.escrow
        escrow_status = escrow.status
        if escrow.commission_amount:
            commission = _dec(escrow.commission_amount)
        if escrow.released_to_driver:
            driver_net = _dec(escrow.released_to_driver)
        elif commission:
            driver_net = amount - commission
        else:
            pct = commission_percent()
            commission = (amount * pct / Decimal('100')).quantize(Decimal('0.01'))
            driver_net = amount - commission
    except ObjectDoesNotExist:
        pct = commission_percent()
        commission = (amount * pct / Decimal('100')).quantize(Decimal('0.01'))
        driver_net = amount - commission

    distance_km = None
    if order.tracked_distance_meters:
        distance_km = round(order.tracked_distance_meters / 1000, 1)
    elif order.optimized_route_distance_meters:
        distance_km = round(order.optimized_route_distance_meters / 1000, 1)

    if pod and pod.receiver_name:
        consignee['name'] = pod.receiver_name

    return {
        'order_id': order.id,
        'status': getattr(order.status, 'code', ''),
        'currency': getattr(ad, 'currency', None) or 'UZS',
        'title': (ad.title_uz or ad.title_ru or ad.title_en) if ad else '',
        'cargo_category': getattr(ad, 'cargo_category', '') or '',
        'weight': str(getattr(ad, 'weight', '') or ''),
        'volume_m3': str(getattr(ad, 'volume_m3', '') or ''),
        'units_count': getattr(ad, 'units_count', None),
        'departure_city': _city_name(dep_city),
        'departure_address': getattr(ad, 'departure_address', '') or '',
        'destination_city': _city_name(dest_city),
        'destination_address': getattr(ad, 'destination_address', '') or '',
        'departure_country': dep_country,
        'destination_country': dest_country,
        'pickup_full': pickup_full,
        'delivery_full': delivery_full,
        'route': route,
        'is_international': bool(dep_country and dest_country and dep_country != dest_country),
        'contact_name': getattr(ad, 'contact_name', '') or '',
        'contact_phone': getattr(ad, 'contact_phone', '') or '',
        'receiver_name': consignee['name'],
        'receiver_phone': consignee['phone'],
        'client': client,
        'driver': driver,
        'shipper': shipper,
        'carrier': carrier,
        'consignee': consignee,
        'company_name': shipper['name'],
        'amount_words': amount_in_words_uz(amount),
        'vehicle': {
            'make': getattr(vehicle, 'make', '') or '',
            'model': getattr(vehicle, 'model', '') or '',
            'number': getattr(vehicle, 'number', '') or '',
        } if vehicle else {'make': '', 'model': '', 'number': ''},
        'stops': stops,
        'pod_receiver': pod.receiver_name if pod else consignee['name'],
        'pod_at': _dt(pod.delivered_at) if pod else '',
        'amount': str(amount),
        'commission': str(commission),
        'driver_net': str(driver_net),
        'escrow_status': escrow_status,
        'created_at': _dt(order.created_at),
        'started_at': _dt(order.started_at),
        'completed_at': _dt(order.completed_at),
        'distance_km': distance_km,
        'generated_at': _dt(timezone.now()),
    }


from apps.orders.document_html import render_document_html
from apps.orders.document_pdf import render_document_pdf


def render_document_xlsx(doc_type: str, snapshot: dict[str, Any], number: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = DOC_TITLES.get(doc_type, 'Hujjat')[:31]
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1D4ED8')
    label_fill = PatternFill('solid', fgColor='F3F4F6')
    thin = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    ws['A1'] = DOC_TITLES.get(doc_type, 'Hujjat')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = number
    ws['B2'] = f"Buyurtma #{snapshot.get('order_id')}"

    client = snapshot.get('client') or {}
    driver = snapshot.get('driver') or {}
    shipper = snapshot.get('shipper') or {}
    carrier = snapshot.get('carrier') or {}
    vehicle = snapshot.get('vehicle') or {}
    rows = [
        ('Sana', snapshot.get('completed_at') or snapshot.get('created_at')),
        ('Hujjat turi', DOC_TITLES.get(doc_type, doc_type)),
        ('Xaridor / jo\'natuvchi', shipper.get('name') or snapshot.get('company_name') or client.get('name')),
        ('STIR', shipper.get('inn') or client.get('inn')),
        ('Manzil', shipper.get('address')),
        ('Bank / hisob', ' '.join(part for part in (shipper.get('bank_name'), shipper.get('bank_account'), shipper.get('mfo')) if part)),
        ('Xaridor telefon', shipper.get('phone') or client.get('phone')),
        ('Tashuvchi', carrier.get('name') or driver.get('name')),
        ('Tashuvchi telefon', carrier.get('phone') or driver.get('phone')),
        ('Transport raqami', vehicle.get('number')),
        ('Yuk', snapshot.get('title')),
        ('Qayerdan', snapshot.get('pickup_full') or f"{snapshot.get('departure_city')} {snapshot.get('departure_address')}".strip()),
        ('Qayerga', snapshot.get('delivery_full') or f"{snapshot.get('destination_city')} {snapshot.get('destination_address')}".strip()),
        ('Og\'irlik kg', snapshot.get('weight')),
        ('Summa', snapshot.get('amount')),
        ("So'z bilan", snapshot.get('amount_words')),
        ('Valyuta', snapshot.get('currency')),
        ('Komissiya', snapshot.get('commission')),
        ('Haydovchiga', snapshot.get('driver_net')),
        ('Qabul qiluvchi', snapshot.get('receiver_name') or snapshot.get('pod_receiver')),
        ('POD vaqti', snapshot.get('pod_at')),
        ('Masofa km', snapshot.get('distance_km')),
    ]
    ws['A4'] = 'Maydon'
    ws['B4'] = 'Qiymat'
    ws['A4'].font = header_font
    ws['B4'].font = header_font
    ws['A4'].fill = header_fill
    ws['B4'].fill = header_fill
    for idx, (label, value) in enumerate(rows, start=5):
        ws[f'A{idx}'] = label
        ws[f'B{idx}'] = '' if value is None else str(value)
        ws[f'A{idx}'].fill = label_fill
        ws[f'A{idx}'].border = thin
        ws[f'B{idx}'].border = thin
        ws[f'B{idx}'].alignment = Alignment(wrap_text=True)

    export = wb.create_sheet('1C')
    export.append([
        'date', 'number', 'order_id', 'inn', 'counterparty', 'route',
        'weight_kg', 'amount', 'commission', 'driver_amount', 'currency', 'doc_type',
    ])
    for cell in export[1]:
        cell.font = header_font
        cell.fill = header_fill
    export.append([
        snapshot.get('completed_at') or snapshot.get('created_at'),
        number,
        snapshot.get('order_id'),
        client.get('inn'),
        snapshot.get('company_name') or client.get('name'),
        f"{snapshot.get('departure_city')} - {snapshot.get('destination_city')}",
        snapshot.get('weight'),
        snapshot.get('amount'),
        snapshot.get('commission'),
        snapshot.get('driver_net'),
        snapshot.get('currency'),
        doc_type,
    ])
    for sheet in (ws, export):
        sheet.column_dimensions['A'].width = 28
        sheet.column_dimensions['B'].width = 48
        for col in range(1, 13):
            sheet.column_dimensions[get_column_letter(col)].width = max(
                sheet.column_dimensions[get_column_letter(col)].width or 12, 14
            )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _save_file(field, filename: str, payload: bytes) -> None:
    if field:
        field.delete(save=False)
    field.save(filename, ContentFile(payload), save=False)


@transaction.atomic
def ensure_order_documents(
    order: Order,
    *,
    doc_types: list[str] | None = None,
    force: bool = False,
) -> list[OrderDocument]:
    from apps.orders.models import OrderDocument

    types = doc_types or [
        TYPE_INVOICE,
        TYPE_TTN,
        TYPE_CMR,
        TYPE_ACT,
    ]
    snapshot = build_document_snapshot(order)
    created: list[OrderDocument] = []
    for doc_type in types:
        if doc_type not in DOC_PREFIX:
            continue
        number = document_number(doc_type, order.id)
        doc, is_new = OrderDocument.objects.get_or_create(
            order=order,
            doc_type=doc_type,
            defaults={'number': number, 'snapshot': snapshot},
        )
        has_pdf = bool(getattr(doc.pdf_file, 'name', ''))
        if not is_new and not force and doc.html_file and doc.xlsx_file and has_pdf:
            created.append(doc)
            continue
        doc.number = number
        doc.snapshot = snapshot
        html_body = render_document_html(doc_type, snapshot, number)
        xlsx_body = render_document_xlsx(doc_type, snapshot, number)
        _save_file(doc.html_file, f'{number}.html', html_body.encode('utf-8'))
        _save_file(doc.xlsx_file, f'{number}.xlsx', xlsx_body)
        try:
            pdf_body = render_document_pdf(doc_type, snapshot, number)
        except Exception:
            logger.exception(
                'Order PDF generation failed',
                extra={'event': 'order_document_pdf_failed', 'order_id': order.id, 'reason': doc_type},
            )
            pdf_body = None
        if pdf_body:
            _save_file(doc.pdf_file, f'{number}.pdf', pdf_body)
        doc.save()
        created.append(doc)
        logger.info(
            'Order document generated',
            extra={'event': 'order_document_generated', 'order_id': order.id, 'reason': doc_type},
        )
    return created


def serialize_order_document(doc: OrderDocument, request=None) -> dict[str, Any]:
    public_path = f'/api/orders/documents/public/{doc.download_token}/'
    html_url = public_path
    pdf_url = f'{public_path}?file=pdf'
    xlsx_url = f'{public_path}?file=xlsx'
    if request is not None:
        html_url = request.build_absolute_uri(html_url)
        pdf_url = request.build_absolute_uri(pdf_url)
        xlsx_url = request.build_absolute_uri(xlsx_url)
    return {
        'id': doc.id,
        'doc_type': doc.doc_type,
        'title': DOC_TITLES.get(doc.doc_type, doc.doc_type),
        'number': doc.number,
        'generated_at': doc.generated_at.isoformat() if doc.generated_at else None,
        'html_url': html_url,
        'pdf_url': pdf_url if bool(getattr(doc.pdf_file, 'name', '')) else html_url,
        'xlsx_url': xlsx_url,
        'has_pdf': bool(getattr(doc.pdf_file, 'name', '')),
        'token': str(doc.download_token),
    }
